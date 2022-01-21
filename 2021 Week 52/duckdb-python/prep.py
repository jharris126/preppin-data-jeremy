from duckdb import connect
from os import path
from pandas import read_excel
from openpyxl import load_workbook

# relative file path using os.path so it will work on any os
file = path.join('..', 'PD 2021 Wk 52 Input.xlsx')

# load workbook object into variable and get list of sheets
wb = load_workbook(filename=file)
sheets = wb.sheetnames

# start connection to duckdb outside for loop so both tables are added to same db
conn = connect()

# loop through sheets and load each to duckdb table
for sheet in sheets:
    table = sheet.replace(' ', '_').lower()
    view = 'df_view'

    conn.register(view, read_excel(file, sheet_name=sheet))
    conn.execute(f'create table {table} as select * from {view}')

# apply sql prep logic from prep.sql file in this same directory
sql = open('prep.sql', 'r').read()

# define output file name/path, sql to execute and export, and duckdb command to execute the sql
outfile = '2021-52-output.csv'
export_sql = f"copy ({sql}) to '{outfile}' with (header 1, delimiter ',');"
conn.execute(export_sql)
